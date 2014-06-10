#!/usr/bin/env python
# -*- coding utf-8 -*-

import os
import itertools
import tempfile
import csv
import datetime as dt
from datetime import timedelta
import errno
import types
import numpy as np; np.seterr(all='raise')
import sys
import time

from functools import partial
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from collections import OrderedDict
from numbers import Number
from decimal import Decimal
from scipy.optimize import newton
from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter

def timer(f):
    def wrapper(*args, **kwargs):
        t = time.time()
        res = f(*args, **kwargs)
        print ("Func %s Execution time: %f sec" % (f.__name__, time.time()-t))
        return res
    return wrapper

class NullDevice():
    """Class for hiding all output"""
    def write(self, s):
        pass

class cached_property(object):
    """
    Descriptor (non-data) for building an attribute on-demand on first use.
    """
    def __init__(self, factory, attr_name=None):
        """
        <attr_name> is the name of the attribute.
        <factory> is called such: factory(instance) to build the attribute.
        """
        self._attr_name = attr_name or factory.__name__
        self._factory = factory

    def __get__(self, instance, owner):
        # Build the attribute.
        attr = self._factory(instance)

        # Cache the value; hide ourselves.
        setattr(instance, self._attr_name, attr)

        return attr


class OrderedDefaultdict(OrderedDict):
    def __init__(self, *args, **kwargs):
        newdefault = None
        newargs = ()
        if args:
            newdefault = args[0]
            if not (newdefault is None or callable(newdefault)):
                raise TypeError('first argument must be callable or None')
            newargs = args[1:]
        self.default_factory = newdefault
        super(self.__class__, self).__init__(*newargs, **kwargs)

    def __missing__ (self, key):
        if self.default_factory is None:
            raise KeyError(key)
        self[key] = value = self.default_factory()
        return value

    def __reduce__(self):  # optional, for pickle support
        args = self.default_factory if self.default_factory else tuple()
        return type(self), args, None, None, self.items()


class memoize(object):
    """cache the return value of a method

    This class is meant to be used as a decorator of methods. The return value
    from a given method invocation will be cached on the instance whose method
    was invoked. All arguments passed to a method decorated with memoize must
    be hashable.

    If a memoized method is invoked directly on its class the result will not
    be cached. Instead the method will be invoked like a static method:
    class Obj(object):
        #@memoize
        def add_to(self, arg):
            return self + arg
    Obj.add_to(1) # not enough arguments
    Obj.add_to(1, 2) # returns 3, result is not cached
    """
    def __init__(self, func):
        self.func = func
    def __get__(self, obj, objtype=None):
        if obj is None:
            return self.func
        return partial(self, obj)
    def __call__(self, *args, **kw):
        obj = args[0]
        try:
            cache = obj.__cache
        except AttributeError:
            cache = obj.__cache = {}
        key = (self.func, args[1:], frozenset(kw.items()))
        try:
            res = cache[key]
        except KeyError:
            res = cache[key] = self.func(*args, **kw)
        return res

def memoized(f):
    memory = {}
    def wrapper(*args, **kwargs):
        key = (tuple(args), hash(tuple(sorted(kwargs.items()))))
        if not key in memory:
            memory[key] = f(*args, **kwargs)
        return memory[key]
    return wrapper

def getMultiplier(_from, _to, step):
    """calculates multiplier for number to make it interger (without decimal part)"""
    digits = []
    for number in [_from, _to, step]:
        pre = Decimal(str(number)) % 1
        digit = len(str(pre))
        digits.append(digit)
    max_digits = max(digits)
    return float(10 ** (max_digits))


def floatRange(_from, _to, step, include=False):
    """Generates a list of floating point values over the Range [start, stop]
       with step size step
    Works fine with floating point representation!
    """
    mult = getMultiplier(_from, _to, step)
    # print mult
    int_from = int(round(_from * mult))
    int_to = int(round(_to * mult))
    int_step = int(round(step * mult))
    # print int_from,int_to,int_step
    if include:
        result = range(int_from, int_to + int_step, int_step)
        result = [r for r in result if r <= int_to]
    else:
        result = range(int_from, int_to, int_step)
    # print result
    float_result = [r / mult for r in result]
    return float_result

def getConfigs(dic):
    """Gets dict ,return   dict with keys not starts with _
    Used to get all attrs from class
    """
    result = {}
    for k, v in dic.items():
        if k.startswith('_') or k == 'self':
            continue
        result[k] = v
    return result

def getDaysNoInMonth(date):
    """return number of days in given month"""
    days = monthrange(date.year, date.month)[1]
    return days

@memoized
def get_list_dates( date_start, date_end):
    duration_days = (date_end - date_start).days
    list_dates = list([(date_start+timedelta(days=i)) for i in range(duration_days+1)])
    return list_dates

def monthsBetween(date1,date2):
    """return full month number since date2 till date1"""
    if date1>date2:
        return (-monthsBetween(date2, date1))
    m1=date1.year*12+date1.month
    m2=date2.year*12+date2.month
    months=m2-m1
    if date1.day>date2.day:
        months-=1
    #elif date1.day==date2.day:

        #seconds1=date1.hour*3600+date1.minute+date1.second
        #seconds2=date2.hour*3600+date2.minute+date2.second
        #if seconds1>seconds2:
            #months-=1
    return months

def isLastDayYear(date):
    """return True is date is last day of year"""
    new_date = date + timedelta(days=1)
    if new_date.year != date.year:
        return True
    else:
        return False

def yearsBetween1Jan(date1,date2):
    """return year difference between 2 dates """
    return date2.year - date1.year

def lastDayMonth(date):
    """return date - last day date in month with input date"""
    day = getDaysNoInMonth(date)
    return dt.date( date.year, date.month, day)

def lastDayYear(date):
    """return date - last day date in month with input date"""
    return dt.date( date.year, 12, 31)

def nubmerDaysInMonth(date):
    """return  number of days in month to which date belongs"""
    last_day = lastDayMonth(date)
    return last_day.day

def sameDayLastYear(date):
    """Return the same date but previous year"""
    last_year = date.year - 1
    last_month = date.month
    temp_date = dt.date(last_year, last_month, 1)
    return  lastDayMonth(temp_date)

def lastDayPrevMonth(date):
    """return date - last day of previous month"""
    prev_month_date =  date - relativedelta(months=1)
    last_day_prev_month = lastDayMonth(prev_month_date)
    return last_day_prev_month

def lastDayNextMonth(date):
    """return date - last day of next month"""
    next_month_date =  date + relativedelta(months=1)
    last_day_next_month = lastDayMonth(next_month_date)
    return last_day_next_month

def firstDayMonth(date):
    """return date - first day date in month with input date"""
    return dt.date( date.year, date.month, 1)

def addXYears(date, years):
    """return date X years later - 1 day"""
    return  date + relativedelta(years=int(years)) - relativedelta(days=1)

def addXMonths(date, months):
    """return date X months later - 1 day"""
    return  date + relativedelta(months=int(months)) - relativedelta(days=1)

@memoized
def getListDates( date_start, date_end):
    """generate list of all dates from date_start to date_end"""
    duration_days = (date_end - date_start).days
    list_dates = list([(date_start+timedelta(days=i)) for i in range(duration_days+1)])
    return list_dates

def convertDictDates(dic):
    """Converts dates from string to object Date
    return like this:
       OrderedDict([(datetime.date(2013, 1, 1), 0.06997119790546502), (datetime.date(2013, 1, 2), 0.07017252445198603),
    """
    return OrderedDict((dt.datetime.strptime(x, '%Y-%m-%d').date(), y) for x, y in sorted(dic.items()))

def uniquifyFilename(path, sep = ''):
    """Return free filename, if file name is Busy adds suffix _number
    Input: filename we want
    Output: filename we can
    """
    def name_sequence():
        count = itertools.count()
        yield ''
        while True:
            yield '{s}_v{n:d}'.format(s = sep, n = next(count))

    orig = tempfile._name_sequence
    with tempfile._once_lock:
        tempfile._name_sequence = name_sequence()
        path = os.path.normpath(path)
        dirname, basename = os.path.split(path)
        filename, ext = os.path.splitext(basename)
        fd, filename = tempfile.mkstemp(dir = dirname, prefix = filename, suffix = ext)
        tempfile._name_sequence = orig
        os.close(fd)
    return filename

def getReportDates(start_date_project, end_date_project):
    """return  dates for reports
    dic_monthly[first_day_month]=last_day_month
    dic_yearly[last_day_month]=list_of_ (start_date,end_date)
    """
    report_dates = OrderedDict()
    report_dates_y = OrderedDefaultdict(list)

    date = firstDayMonth(start_date_project)
    date_to = end_date_project

    while True:
        end_date = lastDayMonth(date)
        end_date_y = lastDayYear(date)
        start_date = firstDayMonth(date)

        report_dates[start_date] = end_date
        report_dates_y[end_date_y].append((start_date, end_date))

        date = addXMonths(date, 1)
        if  date > date_to:
            break

    return (report_dates, report_dates_y)

def transponseCsv(csv_filename):
    """Transposing csv file"""
    new_filenmame = csv_filename
    with  open(csv_filename) as csv_file:
        a = itertools.izip(*csv.reader(csv_file, delimiter=';'))
        with open(new_filenmame, "wb") as f:
            csv.writer(f, delimiter=';').writerows(a)

def addHeaderCsv(csv_filename, header):
    """Adds first row header to csv file"""
    line = ";".join(header)

    with open(csv_filename,'r+') as f:
        content = f.read()
        f.seek(0,0)
        f.write(line.rstrip('\r\n') + '\n' + content)

def mkdir_p(path):
    """ Mkdir if not exists """
    try:
        os.makedirs(path)
    except OSError as exc: # Python >2.5
        if exc.errno == errno.EEXIST:
            pass
        else: raise

def csv2xlsx(inputfilename, outputfilename, sheet_no=0, sheet_name='report'):
    """Converts csv 2 excel"""

    with open(inputfilename) as fin:
        reader = csv.reader(fin, delimiter=';')

        try:
            wb = load_workbook(outputfilename)
        except:
            open(outputfilename, 'ab').close()
            wb = Workbook()

        try:
            ws = wb.worksheets[sheet_no]
            ws.title = sheet_name
        except IndexError:
            ws = wb.create_sheet(sheet_no, sheet_name)



        columns = []
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                column_letter = get_column_letter((column_index + 1))
                if column_letter not in columns:
                    columns.append(column_letter)
                _ceil = ws.cell('%s%s'%(column_letter, (row_index + 1)))
                _ceil.value = cell
                #_ceil.style.font.bold = True

        xlsSetColumnSizes(ws, columns)
        wb.save(filename = outputfilename)


def xlsSetColumnSizes(ws, columns, sizes=(40, 30), others=12):
    """Sets column sizes for ws
    @sizes - list of sizes since 0 column
    others - size for all others columsn
    """
    for col_ind, size in enumerate(sizes):
        col = columns[col_ind]
        ws.column_dimensions[col].width = size

    for column in columns[len(sizes):]:
        ws.column_dimensions[column].width = others

def combineFiles(from_filenames, to_filename):
    """Combining several files to one
    from_filenames=[of filenames], to_filename="filename"
    """
    content = "\n"
    for f in from_filenames:
        content += open(f).read() + '\n\n\n'
        os.remove(f)

    with open(to_filename,'wb') as output:
        output.write(content)


def addSecondSheetXls(filename, secondsheet_csv):

    csv2xlsx(secondsheet_csv, filename, sheet_name='source', sheet_no=1)
    os.remove(secondsheet_csv)

def convert2excel(source, output):
    """Converts souce file to excel file with name = @output"""
    csv2xlsx(source, output)
    os.remove(source)
    return  output

def getInputDate(default=None, text=''):
    """Reads input date from user"""
    date = raw_input("Input date %s or press ENTER to use default value %s::  " %(text, default))
    try:
        result= dt.datetime.strptime(date,"%Y-%m-%d").date()
    except ValueError:
        print "No value or value error. Return default value %s" % default
        if isinstance(default, (str, unicode)):
            result = dt.datetime.strptime(default, "%Y-%m-%d").date()
        else:
            result = default
    return result

def getInputInt(text='', default=False):
    """reads input integer value from user"""
    value = raw_input(text)
    try:
        result= int(value)
    except ValueError:
        if default is not False:
            print "No value or value error. Return default value %s" % default
            result = default
        else:
            raise ValueError("No value or value error")

    return result

def getInputComment(text='Please input comment for current simulation: ', default='w/o comment' ):
    """reads comment text from user input"""
    value = raw_input(text)
    result= str(value)
    if not result:
        result = default

    return result

def getOnlyDigits(obj):
    """gets only digit values from dict obj"""
    new_obj = obj.values()
    return  getOnlyDigitsList(new_obj)

def getOnlyDigitsList(obj):
    """filters and gets only digit list values"""
    return  filter(lambda x :isinstance(x, Number), obj)

def getResolutionStartEnd(start_date, end_date, resolution):
    """Returns list [start , start_date + resolution] using resolution (days interval)"""
    result = []
    while start_date < end_date:
        next_step_date = start_date + timedelta(days=resolution-1)
        if next_step_date <= end_date:
            result.append((start_date, next_step_date))
        else:
            result.append((start_date, end_date))
        start_date = next_step_date + timedelta(days=1)
    return result


def addYearlyPrefix(field, yearly):
    """adds yearly prefix to field if bool @yearly"""
    if yearly:
        if field.startswith('iterations.'):
            field += '_y'
            return field
        elif field.endswith('_y'):
            return field
        elif field.find('.') == -1:
            field += '_y'
            return field

    return field

class Annuitet():
    def __init__(self,summa,yrate,yperiods, start_date):
        """
        summa - how much we borrow
        yrate - fix rate in percents per year
        yperiods - number of years to return debt
        """
        self.summa = summa
        self.yrate = yrate
        self.mperiods = yperiods * 12
        self.mpayment = self.__calcMonthlyPayment()
        self.start_date = start_date
        self.total_debt = self.mpayment * self.mperiods


    def __calcMonthlyPayment(self):
        P = self.yrate / 12
        N = self.mperiods
        S = self.summa

        return  S * P * (1+P) ** N / ((1+P) ** N-1)


    def calculate(self):
        """Calculated annuitet values

        @percent_payments = dict with sum we need to pay at current period - for percent of loan
        @debt_payments = dict with sum we need to pay at current period - for body of loan
        @rest_payments = dict with sum, we still need to pay in next periods

        """
        date = self.start_date

        percent_payment = 0# (self.summa * self.yrate / 12)
        debt_payment = 0
        rest_payment = self.total_debt
        rest_payment_wo_percent = self.summa

        percent_payments = OrderedDict({date: percent_payment,})
        debt_payments = OrderedDict({date: debt_payment,})
        rest_payments = OrderedDict({date: rest_payment,})
        rest_payments_wo_percents = OrderedDict({date: rest_payment_wo_percent,})

        for i in range(0, self.mperiods):
            date = lastDayNextMonth(date)

            percent_payment = rest_payment_wo_percent * self.yrate / 12
            debt_payment = self.mpayment - percent_payment
            rest_payment -= self.mpayment
            rest_payment_wo_percent -= debt_payment

            if rest_payment < 0.01:
                rest_payment = 0

            if rest_payment_wo_percent < 0.01:
                rest_payment_wo_percent = 0

            percent_payments[date] = percent_payment
            debt_payments[date] = debt_payment
            rest_payments[date] = rest_payment
            rest_payments_wo_percents[date] = rest_payment_wo_percent

        self.percent_payments = percent_payments
        self.debt_payments = debt_payments
        self.rest_payments = rest_payments
        self.rest_payments_wo_percents = rest_payments_wo_percents

def convertDict(input_dict):
    """part of deep_serilization of DICTS"""
    result = OrderedDict()
    for (key, value) in input_dict.items():
        key = convertBaseValues(key)
        value = convertBaseValues(value)
        value = convertValue(value)
        result[key] = value
    return result

def convertBaseValues(base_value):
    """part of deep_serilization of BASE VALUES"""
    if isinstance(base_value, dt.date):
        key = str(base_value)
    else:
        key = base_value
    return key

def convertValue(value):
    """part of deep_serilization MAIN"""
    if isinstance(value, types.DictionaryType):
        return  convertDict(value)
    elif isinstance(value, (types.ListType, types.TupleType)):
        return convertList(value)
    elif isinstance(value, dt.date):
        return str(value)
    else:
        return value

def convertList(input_lst):
    """part of deep_serilization of LISTS and TUPLES"""
    return [convertValue(v) for v in input_lst]

def irr(pmts, guess=0.01):
    """
    IRR function that accepts irregularly spaced cash flows
    @values: array_like
          Contains the cash flows including the initial investment
    @Returns: Float
          Internal Rate of Return
    """
    def _discf(rate, pmts, ):
        """used for irr calculation"""
        dcf=[]
        for i,cf in enumerate(pmts):
            dcf.append(cf*(1+rate)**(-i))
        return np.add.reduce(dcf)

    f = lambda x: _discf(x, pmts)

    try:
        sys.stderr = NullDevice()
        value = newton(f, guess, maxiter=100, tol=10**(-10))
        sys.stderr = sys.__stderr__
        if value > 1:
            print "Too large IRR %s . Setting to zero" % value
            value = 0
        if value < 0.0001:
            print "Too low IRR %s . Setting to zero" % value
            value = 0
        return value
    except RuntimeError:
        return float('Nan')


def setupPrintProgress(format, function=lambda x: x):
    """Prepares function for later printing like progress bar
    format - '%s' like this format for printing
    function - that will use convert step to printed value
    """
    upd_format = '\r' + format

    def func(step=None, stop=False):
        if stop:
            sys.stdout.write('Done\n')
            sys.stdout.flush()
        else:
            value = function(step)
            sys.stdout.write(upd_format % value)
            sys.stdout.flush()

    return func


def print_separator():
    print "=" * 70

if __name__ == '__main__':
    start_date = dt.date(2000, 1, 1)
    end_date = dt.date(2001, 12, 31)
    print getResolutionStartEnd(start_date, end_date, 10)

